#!/usr/bin/env python3
"""extract_bd_custominputs.py

Lee la BD master (XLSM con hoja 'Upload', headers row 7, datos desde row 9) y
extrae el `customInputs` por PN, con la misma estructura que SH espera. Es el
paso 1 del recovery post-blanqueo de la corrida v104.

Output: bd_custominputs.json
    [
      {
        "pn": "...",
        "customer": "...",
        "customInputs": {
          "QuoteIBMS": "...",
          "EstacionIBMS": "...",
          "DatosAdicionalesNP": {"BaseMetal": "...", "Plano": "..."},
          "DatosFacturacion": {"CodigoSAT": "..."},
          "DatosPlanificacion": {
              "PiezasPorCarga": 12, "CargasPorHora": 0.5, "TiempoEntrega": "..."
          },
          "NotasAdicionales": "..."
        }
      },
      ...
    ]

Uso:
  python3 tools/extract_bd_custominputs.py \\
      --bd-xlsm "/Users/oviazcan/Downloads/BD Numeros de Parte Reloaded v23 ...xlsm" \\
      --out bd_custominputs.json
"""
import argparse
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Falta openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# Columnas 1-based en la hoja Upload (row 7 headers)
COL_CLIENTE       = 5
COL_PN            = 6
COL_METAL_BASE    = 15
COL_CODIGO_SAT    = 53
COL_QUOTE_IBMS    = 63
COL_ESTACION_IBMS = 64
COL_PLANO         = 65
COL_PZAS_CARGA    = 66
COL_CARGAS_HORA   = 67
COL_TIEMPO_ENTREGA = 68
COL_NOTAS_ADIC    = 69


def cell(row, col_1based):
    idx = col_1based - 1
    if idx >= len(row):
        return None
    return row[idx]


def s(v):
    if v is None:
        return None
    if isinstance(v, str):
        t = v.strip()
        return t if t else None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip() or None


def n(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, str):
        t = v.strip()
        if not t:
            return None
        try:
            f = float(t)
            return int(f) if f.is_integer() else f
        except ValueError:
            return None
    return None


def build_ci(row):
    ci = {}
    qibms = s(cell(row, COL_QUOTE_IBMS))
    estIbms = s(cell(row, COL_ESTACION_IBMS))
    if qibms: ci["QuoteIBMS"] = qibms
    if estIbms: ci["EstacionIBMS"] = estIbms

    base = s(cell(row, COL_METAL_BASE))
    plano = s(cell(row, COL_PLANO))
    datos_adic = {}
    if base: datos_adic["BaseMetal"] = base
    if plano: datos_adic["Plano"] = plano
    if datos_adic: ci["DatosAdicionalesNP"] = datos_adic

    sat = s(cell(row, COL_CODIGO_SAT))
    if sat:
        ci["DatosFacturacion"] = {"CodigoSAT": sat}

    pcarga = n(cell(row, COL_PZAS_CARGA))
    cph = n(cell(row, COL_CARGAS_HORA))
    te = s(cell(row, COL_TIEMPO_ENTREGA))
    plan = {}
    if pcarga is not None: plan["PiezasPorCarga"] = pcarga
    if cph is not None: plan["CargasPorHora"] = cph
    if te: plan["TiempoEntrega"] = te
    if plan: ci["DatosPlanificacion"] = plan

    notas = s(cell(row, COL_NOTAS_ADIC))
    if notas: ci["NotasAdicionales"] = notas

    return ci if ci else None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--bd-xlsm", required=True)
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    p = Path(args.bd_xlsm)
    if not p.exists():
        print(f"No existe: {p}", file=sys.stderr); sys.exit(2)

    wb = openpyxl.load_workbook(p, read_only=True, data_only=True, keep_links=False)
    if "Upload" not in wb.sheetnames:
        print(f"{p}: sin hoja 'Upload'", file=sys.stderr); sys.exit(2)
    ws = wb["Upload"]

    out = []
    seen_pns = set()
    dups = []
    for r_idx, row in enumerate(ws.iter_rows(min_row=9, values_only=True), start=9):
        pn = s(cell(row, COL_PN))
        if not pn:
            continue
        cli = s(cell(row, COL_CLIENTE))
        ci = build_ci(row)
        if not ci:
            continue
        key = (pn, cli or "")
        if key in seen_pns:
            dups.append({"pn": pn, "customer": cli, "row": r_idx})
            continue
        seen_pns.add(key)
        out.append({
            "pn": pn,
            "customer": cli,
            "row": r_idx,
            "customInputs": ci,
        })

    Path(args.out).write_text(json.dumps(out, ensure_ascii=False, indent=2))
    print(f"✅ {len(out)} PNs con customInputs extraídos → {args.out}")
    if dups:
        print(f"⚠️  {len(dups)} duplicados (pn,customer) — primer ocurrencia conservada")

    # Cobertura
    counts = {
        "QuoteIBMS": 0, "EstacionIBMS": 0, "BaseMetal": 0, "Plano": 0,
        "CodigoSAT": 0, "PiezasPorCarga": 0, "CargasPorHora": 0,
        "TiempoEntrega": 0, "NotasAdicionales": 0,
    }
    for r in out:
        ci = r["customInputs"]
        if ci.get("QuoteIBMS"): counts["QuoteIBMS"] += 1
        if ci.get("EstacionIBMS"): counts["EstacionIBMS"] += 1
        da = ci.get("DatosAdicionalesNP", {})
        if da.get("BaseMetal"): counts["BaseMetal"] += 1
        if da.get("Plano"): counts["Plano"] += 1
        df = ci.get("DatosFacturacion", {})
        if df.get("CodigoSAT"): counts["CodigoSAT"] += 1
        dp = ci.get("DatosPlanificacion", {})
        if dp.get("PiezasPorCarga") is not None: counts["PiezasPorCarga"] += 1
        if dp.get("CargasPorHora") is not None: counts["CargasPorHora"] += 1
        if dp.get("TiempoEntrega"): counts["TiempoEntrega"] += 1
        if ci.get("NotasAdicionales"): counts["NotasAdicionales"] += 1
    print("Cobertura por campo:")
    for k, v in counts.items():
        print(f"  {k:18s}: {v}")


if __name__ == "__main__":
    main()
