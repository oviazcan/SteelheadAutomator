"""Tests for tools/dual_source_recovery.py."""

from __future__ import annotations

import json
import pytest

import openpyxl as _openpyxl

from tools.dual_source_recovery import (
    ROUND_MARKER_TOKENS,
    TYPE_NUMBER,
    TYPE_STRING,
    PartNumberRow,
    compare_values,
    compute_field_diffs,
    emit_v11_xlsx,
    filter_round,
    is_round_marker,
    load_sh_report,
    emit_json_report,
    load_xlsm_originals,
    main,
    match_xlsm_to_sh,
    make_fingerprint,
    norm,
    read_header_map,
    validate_notas,
)


def test_module_imports():
    """El módulo debe importarse sin errores."""
    from tools import dual_source_recovery as m
    assert callable(m.main)


class TestNorm:
    def test_empty_returns_empty(self):
        assert norm("") == ""
        assert norm(None) == ""

    def test_strips_whitespace(self):
        assert norm("  hola  ") == "hola"

    def test_lowercases(self):
        assert norm("HOLA") == "hola"

    def test_collapses_internal_whitespace(self):
        assert norm("a   b\t\nc") == "a b c"

    def test_preserves_special_chars(self):
        assert norm("PLATA 2µm") == "plata 2µm"

    def test_handles_newlines(self):
        assert norm("line1\nline2") == "line1 line2"

    def test_non_string_to_empty(self):
        assert norm(123) == "123"  # cast a string primero
        assert norm(0) == "0"


class TestIsRoundMarker:
    def test_empty_returns_false(self):
        assert is_round_marker("") is False
        assert is_round_marker(None) is False

    def test_pattern_with_5_tokens(self):
        notas = "F1: PLATA | F2: ANTITARNISH | SPECS: PLATA COLGADO | DEPT: 16.3 | METAL: COBRE | PROC: PLATA SELECTIVA"
        assert is_round_marker(notas) is True

    def test_pattern_with_mpo(self):
        notas = "F1: Decapado | F2: ESTAÑO | MPO: DECAPADO ESTAÑADO"
        assert is_round_marker(notas) is True  # F1, F2, MPO = 3 tokens

    def test_only_2_tokens_returns_false(self):
        notas = "F1: PLATA | F2: ANTITARNISH"
        assert is_round_marker(notas) is False  # menos de 3

    def test_no_pipe_separator_returns_false(self):
        notas = "F1 PLATA SPECS COBRE DEPT 16.3"
        assert is_round_marker(notas) is False

    def test_freeform_text_returns_false(self):
        assert is_round_marker("EMPAQUE CON PAPEL SILVER SAVER") is False
        assert is_round_marker("Plata Flash Conector") is False

    def test_case_insensitive_tokens(self):
        # los tokens son case-sensitive en SH (F1:, no f1:), pero la robustez no estorba
        notas = "f1: plata | specs: plata | dept: 16 | metal: cobre"
        assert is_round_marker(notas) is True

    def test_tokens_constant_exists(self):
        assert "F1:" in ROUND_MARKER_TOKENS
        assert "MPO:" in ROUND_MARKER_TOKENS
        assert "SPECS:" in ROUND_MARKER_TOKENS


class TestMakeFingerprint:
    def test_basic(self):
        fp = make_fingerprint(metal_base="COBRE", labels=["PLATA", "ANTITARNISH"])
        assert fp == "COBRE|ANTITARNISH,PLATA"  # labels sorted

    def test_normalizes_case_and_spaces(self):
        fp1 = make_fingerprint(metal_base="cobre", labels=[" PLATA ", "antitarnish"])
        fp2 = make_fingerprint(metal_base="COBRE", labels=["ANTITARNISH", "PLATA"])
        assert fp1 == fp2

    def test_drops_empty_labels(self):
        fp = make_fingerprint(metal_base="COBRE", labels=["PLATA", "", None, "ANTITARNISH"])
        assert fp == "COBRE|ANTITARNISH,PLATA"

    def test_no_labels(self):
        fp = make_fingerprint(metal_base="ACERO", labels=[])
        assert fp == "ACERO|"

    def test_no_metal(self):
        fp = make_fingerprint(metal_base="", labels=["PLATA"])
        assert fp == "|PLATA"


class TestReadHeaderMap:
    def test_maps_headers_to_indices(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test"
        ws["A1"] = "Cliente"
        ws["B1"] = "Número de parte"
        ws["C1"] = "Etiqueta 1"
        path = tmp_path / "test.xlsx"
        wb.save(path)

        wb2 = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws2 = wb2["Test"]
        m = read_header_map(ws2, header_row=1)
        assert m == {"Cliente": 1, "Número de parte": 2, "Etiqueta 1": 3}

    def test_normalizes_newlines_in_headers(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Validación\n1er recibo"
        ws["B1"] = "Esp. Spec 1\n(µm)"
        path = tmp_path / "test.xlsx"
        wb.save(path)

        wb2 = openpyxl.load_workbook(path, read_only=True, data_only=True)
        m = read_header_map(wb2.active, header_row=1)
        assert m == {"Validación 1er recibo": 1, "Esp. Spec 1 (µm)": 2}

    def test_skips_empty_columns(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "Cliente"
        ws["C1"] = "Proceso"  # B1 empty
        path = tmp_path / "test.xlsx"
        wb.save(path)

        wb2 = openpyxl.load_workbook(path, read_only=True, data_only=True)
        m = read_header_map(wb2.active, header_row=1)
        assert m == {"Cliente": 1, "Proceso": 3}

    def test_header_row_param(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A7"] = "Cliente"
        ws["B7"] = "Proceso"
        path = tmp_path / "test.xlsx"
        wb.save(path)

        wb2 = openpyxl.load_workbook(path, read_only=True, data_only=True)
        m = read_header_map(wb2.active, header_row=7)
        assert m == {"Cliente": 1, "Proceso": 2}


def _make_sh_report(tmp_path, rows):
    """Crea un xlsx tipo 'reporte SH' con headers en row 1 y rows desde row 2.
    `rows` es lista de dicts {header → value}.
    """
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ING. Validación de Carga Masiva"
    headers = [
        "Archivado", "Validación\n1er recibo", "Forzar\nduplicar", "Archivar\nanterior",
        "Id SH", "Cliente", "Número de\nparte", "Descripción", "PN alterno", "Grupo",
        "Cantidad", "Precio", "Unidad\nprecio", "Divisa", "Precio\ndefault",
        "Línea", "Metal base",
        "Etiqueta 1", "Etiqueta 2", "Etiqueta 3", "Etiqueta 4", "Etiqueta 5",
        "Proceso",
        "Producto 1", "Precio P1", "Cant P1", "Unidad P1",
        "Producto 2", "Precio P2", "Cant P2", "Unidad P2",
        "Producto 3", "Precio P3", "Cant P3", "Unidad P3",
        "Spec 1", "Esp. Spec 1\n(µm)", "Spec 2", "Esp. Spec 2\n(µm)",
        "KGM\n(kg/pza)", "CMK\n(cm²/pza)", "LM\n(m/pza)", "Mín Pzas\nLote",
        "Rack Flybar o Barril (Carga)", "Pzas/Rack\nLínea", "Rack Específico", "Pzas/Rack\nSec.",
        "Tipo de Geometría", "Longitud\n(m)", "Ancho\n(m)", "Alto\n(m)", "Diám.Ext\n(m)", "Diám.Int\n(m)",
        "Departamento", "Código SAT",
        "Plata\n(kg/pza)", "Estaño\n(kg/pza)", "Níquel\n(kg/pza)", "Zinc\n(kg/pza)", "Cobre\n(kg/pza)",
        "Antitarnish\n(L/pza)", "Epóx. MT\n(lb/pza)", "Epóx. BT\n(lb/pza)", "Epóx. MTR\n(lb/pza)",
        "Notas adicionales", "QuoteIBMS", "EstIBMS", "Plano",
        "Piezas por Carga", "Cargas por Hora", "Tiempo de Entrega",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    # normalizar cada dict de row contra los headers limpios (sin \n)
    import re as _re
    clean = {_re.sub(r"\s+", " ", h.replace("\n", " ").strip()): i for i, h in enumerate(headers, start=1)}
    for r_idx, row_dict in enumerate(rows, start=2):
        for key, val in row_dict.items():
            col = clean.get(key)
            if col:
                ws.cell(row=r_idx, column=col, value=val)
    path = tmp_path / "sh_report.xlsx"
    wb.save(path)
    return path


class TestLoadShReport:
    def test_loads_basic_row(self, tmp_path):
        path = _make_sh_report(tmp_path, [
            {"Id SH": 12345, "Cliente": "SCHNEIDER", "Número de parte": "ABC-001",
             "Descripción": "TEST", "Metal base": "COBRE",
             "Etiqueta 1": "PLATA", "Etiqueta 2": "ANTITARNISH",
             "Proceso": "PLATA SELECTIVA",
             "Notas adicionales": "F1: PLATA | F2: ANTITARNISH | SPECS: PLATA | DEPT: 16 | METAL: COBRE | PROC: PLATA",
             "QuoteIBMS": "Q-001"},
        ])
        rows = load_sh_report(path)
        assert len(rows) == 1
        r = rows[0]
        assert r.id_sh == "12345"
        assert r.cliente == "SCHNEIDER"
        assert r.pn == "ABC-001"
        assert r.metal_base == "COBRE"
        assert r.labels == ["PLATA", "ANTITARNISH", "", "", ""]
        assert r.proceso == "PLATA SELECTIVA"
        assert r.quote_ibms == "Q-001"
        assert r.source == "sh_report"
        assert r.source_row == 2

    def test_skips_completely_empty_rows(self, tmp_path):
        path = _make_sh_report(tmp_path, [
            {"Id SH": 1, "Cliente": "X", "Número de parte": "A"},
            {},  # empty row
            {"Id SH": 2, "Cliente": "Y", "Número de parte": "B"},
        ])
        rows = load_sh_report(path)
        assert len(rows) == 2
        assert {r.pn for r in rows} == {"A", "B"}

    def test_raw_dict_includes_all_headers(self, tmp_path):
        path = _make_sh_report(tmp_path, [
            {"Id SH": 1, "Cliente": "X", "Número de parte": "A",
             "Plata (kg/pza)": 0.5, "Notas adicionales": "F1: X | F2: Y | SPECS: Z"},
        ])
        rows = load_sh_report(path)
        assert rows[0].raw["Plata (kg/pza)"] == 0.5
        assert rows[0].raw["Notas adicionales"] == "F1: X | F2: Y | SPECS: Z"

    def test_aborts_if_required_header_missing(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ING. Validación de Carga Masiva"
        ws["A1"] = "Cliente"  # falta Id SH, Número de parte, etc.
        path = tmp_path / "bad.xlsx"
        wb.save(path)

        with pytest.raises(ValueError, match="header esperado"):
            load_sh_report(path)


def _make_xlsm(tmp_path, name, rows):
    """Crea un xlsm tipo 'BD Numeros de Parte' con headers en row 7 + datos desde row 9.

    `rows` es lista de dicts {header → value}.
    """
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Upload"
    # Row 7 headers (subset suficiente para tests)
    headers = [
        "Archivado", "Validación\n1er recibo", "Forzar\nduplicar", "Archivar\nanterior",
        "Cliente", "Número de\nparte", "Descripción", "PN alterno", "Grupo",
        "Cantidad", "Precio", "Unidad\nprecio", "Divisa", "Precio\ndefault",
        "Metal base",
        "Etiqueta 1", "Etiqueta 2", "Etiqueta 3", "Etiqueta 4", "Etiqueta 5",
        "Proceso",
        "Producto 1", "Precio P1", "Cant P1", "Unidad P1",
        "Producto 2", "Precio P2", "Cant P2", "Unidad P2",
        "Producto 3", "Precio P3", "Cant P3", "Unidad P3",
        "Spec 1", "Esp. Spec 1\n(µm)", "Spec 2", "Esp. Spec 2\n(µm)",
        "KGM\n(kg/pza)", "CMK\n(cm²/pza)", "LM\n(m/pza)", "Mín Pzas\nLote",
        "Rack Flybar o Barril (Carga)", "Pzas/Rack\nLínea", "Rack Específico", "Pzas/Rack\nSec.",
        "Longitud\n(m)", "Ancho\n(m)", "Alto\n(m)", "Diám.Ext\n(m)", "Diám.Int\n(m)",
        "Línea", "Departamento", "Código SAT",
        "Plata\n(kg/pza)", "Estaño\n(kg/pza)", "Níquel\n(kg/pza)", "Zinc\n(kg/pza)", "Cobre\n(kg/pza)",
        "Antitarnish\n(L/pza)", "Epóx. MT\n(lb/pza)", "Epóx. BT\n(lb/pza)", "Epóx. MTR\n(lb/pza)",
        "Notas adicionales", "QuoteIBMS", "EstIBMS", "Plano",
        "Piezas por Carga", "Cargas por Hora", "Tiempo de Entrega",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=7, column=i, value=h)
    # row 8 subheaders (V/F, Texto, etc.) — el script los ignora
    ws.cell(row=8, column=1, value="V/F")
    import re as _re
    clean = {_re.sub(r"\s+", " ", h.replace("\n", " ").strip()): i for i, h in enumerate(headers, start=1)}
    for r_idx, row_dict in enumerate(rows, start=9):
        for key, val in row_dict.items():
            col = clean.get(key)
            if col:
                ws.cell(row=r_idx, column=col, value=val)
    path = tmp_path / name
    wb.save(path)
    return path


class TestLoadXlsmOriginals:
    def test_loads_basic_row(self, tmp_path):
        path = _make_xlsm(tmp_path, "test.xlsm", [
            {"Cliente": "SCHNEIDER", "Número de parte": "ABC-001",
             "Metal base": "COBRE", "Etiqueta 1": "PLATA",
             "Proceso": "PLATA SELECTIVA",
             "Notas adicionales": "F1: PLATA | SPECS: PLATA | DEPT: 16 | METAL: COBRE | PROC: PLATA",
             "QuoteIBMS": "Q-001"},
        ])
        rows = load_xlsm_originals([(path, "xlsm_test")])
        assert len(rows) == 1
        r = rows[0]
        assert r.cliente == "SCHNEIDER"
        assert r.pn == "ABC-001"
        assert r.metal_base == "COBRE"
        assert r.labels[0] == "PLATA"
        assert r.proceso == "PLATA SELECTIVA"
        assert r.quote_ibms == "Q-001"
        assert r.source == "xlsm_test"
        assert r.id_sh == ""  # xlsm no tiene Id SH

    def test_multiple_files_concatenates(self, tmp_path):
        p1 = _make_xlsm(tmp_path, "a.xlsm", [
            {"Cliente": "C1", "Número de parte": "PN1"},
        ])
        p2 = _make_xlsm(tmp_path, "b.xlsm", [
            {"Cliente": "C2", "Número de parte": "PN2"},
            {"Cliente": "C3", "Número de parte": "PN3"},
        ])
        rows = load_xlsm_originals([(p1, "xlsm_srg"), (p2, "xlsm_cg")])
        assert len(rows) == 3
        assert [r.source for r in rows].count("xlsm_srg") == 1
        assert [r.source for r in rows].count("xlsm_cg") == 2

    def test_skips_rows_without_pn(self, tmp_path):
        path = _make_xlsm(tmp_path, "test.xlsm", [
            {"Cliente": "C1", "Número de parte": "PN1"},
            {"Cliente": "C2"},  # no PN
            {},  # vacío
        ])
        rows = load_xlsm_originals([(path, "xlsm_test")])
        assert len(rows) == 1


class TestFilterRound:
    def test_filters_by_marker(self, tmp_path):
        path = _make_sh_report(tmp_path, [
            {"Id SH": 1, "Cliente": "C", "Número de parte": "A",
             "Notas adicionales": "F1: X | F2: Y | SPECS: Z | DEPT: 1 | METAL: M"},
            {"Id SH": 2, "Cliente": "C", "Número de parte": "B",
             "Notas adicionales": "EMPAQUE CON PAPEL SILVER SAVER"},
            {"Id SH": 3, "Cliente": "C", "Número de parte": "C",
             "Notas adicionales": ""},
        ])
        rows = load_sh_report(path)
        filtered = filter_round(rows)
        assert len(filtered) == 1
        assert filtered[0].id_sh == "1"


def _mk_row(source="sh_report", id_sh="", cliente="C", pn="P", quote_ibms="",
            metal_base="", labels=None, notas=""):
    return PartNumberRow(
        source=source, source_row=2, id_sh=id_sh, cliente=cliente, pn=pn,
        descripcion="", quote_ibms=quote_ibms, est_ibms="", notas=notas,
        metal_base=metal_base, labels=labels or ["", "", "", "", ""],
        proceso="", spec1="", spec1_um="", spec2="", spec2_um="", raw={},
    )


class TestMatchXlsmToSh:
    def test_tier1_quoteibms(self):
        xlsm = [_mk_row(source="xlsm", quote_ibms="Q-001", cliente="C1", pn="A")]
        sh = [_mk_row(id_sh="100", quote_ibms="Q-001", cliente="C1", pn="A")]
        result = match_xlsm_to_sh(xlsm, sh)
        assert len(result.matched) == 1
        assert result.matched[0].tier == "quoteIBMS"
        assert result.matched[0].sh_row.id_sh == "100"
        assert result.unmatched_xlsm == []
        assert result.duplicate_quoteibms == []

    def test_tier1_duplicate_quoteibms_falls_to_tier2(self):
        xlsm = [_mk_row(source="xlsm", quote_ibms="Q-001", cliente="C1", pn="A")]
        sh = [
            _mk_row(id_sh="100", quote_ibms="Q-001", cliente="C1", pn="A"),
            _mk_row(id_sh="200", quote_ibms="Q-001", cliente="C2", pn="B"),
        ]
        result = match_xlsm_to_sh(xlsm, sh)
        # 2 candidatos por QuoteIBMS → duplicate bucket + cae a tier 2
        assert len(result.duplicate_quoteibms) == 1
        assert result.duplicate_quoteibms[0]["quoteIBMS"] == "Q-001"
        # tier 2 resuelve por (C1, A) único → match con id_sh=100
        assert len(result.matched) == 1
        assert result.matched[0].tier == "composite_unique"
        assert result.matched[0].sh_row.id_sh == "100"

    def test_tier2_composite_unique(self):
        xlsm = [_mk_row(source="xlsm", cliente="C1", pn="A")]
        sh = [_mk_row(id_sh="100", cliente="C1", pn="A")]
        result = match_xlsm_to_sh(xlsm, sh)
        assert len(result.matched) == 1
        assert result.matched[0].tier == "composite_unique"

    def test_tier3_fingerprint(self):
        xlsm = [_mk_row(source="xlsm", cliente="C1", pn="A",
                        metal_base="COBRE", labels=["PLATA", "", "", "", ""])]
        sh = [
            _mk_row(id_sh="100", cliente="C1", pn="A",
                    metal_base="COBRE", labels=["PLATA", "", "", "", ""]),
            _mk_row(id_sh="200", cliente="C1", pn="A",
                    metal_base="ACERO", labels=["ZINC", "", "", "", ""]),
        ]
        result = match_xlsm_to_sh(xlsm, sh)
        assert len(result.matched) == 1
        assert result.matched[0].tier == "fingerprint"
        assert result.matched[0].sh_row.id_sh == "100"

    def test_unmatched_no_pn_in_sh_round(self):
        xlsm = [_mk_row(source="xlsm", cliente="C1", pn="A")]
        sh = [_mk_row(id_sh="100", cliente="OTHER", pn="OTHER")]
        result = match_xlsm_to_sh(xlsm, sh)
        assert result.matched == []
        assert len(result.unmatched_xlsm) == 1
        assert result.unmatched_xlsm[0]["reason"] == "no_pn_in_sh_round"

    def test_unmatched_ambiguous_fingerprint(self):
        xlsm = [_mk_row(source="xlsm", cliente="C1", pn="A",
                        metal_base="COBRE", labels=["PLATA", "", "", "", ""])]
        sh = [
            _mk_row(id_sh="100", cliente="C1", pn="A",
                    metal_base="COBRE", labels=["PLATA", "", "", "", ""]),
            _mk_row(id_sh="200", cliente="C1", pn="A",
                    metal_base="COBRE", labels=["PLATA", "", "", "", ""]),
        ]
        result = match_xlsm_to_sh(xlsm, sh)
        assert result.matched == []
        assert len(result.unmatched_xlsm) == 1
        assert result.unmatched_xlsm[0]["reason"] == "ambiguous_fingerprint"

    def test_unmatched_sh_in_round(self):
        xlsm = [_mk_row(source="xlsm", cliente="C1", pn="A")]
        sh = [
            _mk_row(id_sh="100", cliente="C1", pn="A"),
            _mk_row(id_sh="200", cliente="C2", pn="B"),
        ]
        result = match_xlsm_to_sh(xlsm, sh)
        assert len(result.unmatched_sh_in_round) == 1
        assert result.unmatched_sh_in_round[0]["idSH"] == "200"


class TestValidateNotas:
    def test_both_empty_returns_ok(self):
        x = _mk_row(notas="")
        s = _mk_row(notas="")
        assert validate_notas(x, s) == "ok"

    def test_both_equal_returns_ok(self):
        x = _mk_row(notas="F1: PLATA | SPECS: PLATA | DEPT: 16")
        s = _mk_row(notas="F1: PLATA | SPECS: PLATA | DEPT: 16")
        assert validate_notas(x, s) == "ok"

    def test_normalizes_whitespace(self):
        x = _mk_row(notas="F1: PLATA  |  SPECS: PLATA")
        s = _mk_row(notas="f1:    plata | specs: plata")
        assert validate_notas(x, s) == "ok"

    def test_sh_empty_xlsm_full_returns_ok(self):
        # carga parcial: el PN se creó sin notas pero el match es válido
        x = _mk_row(notas="F1: PLATA | SPECS: PLATA")
        s = _mk_row(notas="")
        assert validate_notas(x, s) == "ok"

    def test_both_full_but_different_returns_suspicious(self):
        x = _mk_row(notas="F1: PLATA | SPECS: PLATA | DEPT: 16")
        s = _mk_row(notas="F1: ZINC | SPECS: ZINC | DEPT: 7")
        assert validate_notas(x, s) == "suspicious"

    def test_xlsm_empty_sh_full_returns_ok(self):
        # xlsm sin notas: no podemos validar, confiar en el tier
        x = _mk_row(notas="")
        s = _mk_row(notas="F1: PLATA | SPECS: PLATA")
        assert validate_notas(x, s) == "ok"


class TestCompareValues:
    def test_strings_equal_normalized(self):
        assert compare_values("PLATA", "plata", TYPE_STRING) is True
        assert compare_values("  PLATA  ", "plata", TYPE_STRING) is True

    def test_strings_different(self):
        assert compare_values("PLATA", "ZINC", TYPE_STRING) is False

    def test_numbers_within_tolerance(self):
        assert compare_values(0.500001, 0.5, TYPE_NUMBER) is True
        assert compare_values("0.5", 0.5, TYPE_NUMBER) is True

    def test_numbers_different(self):
        assert compare_values(0.5, 0.6, TYPE_NUMBER) is False

    def test_one_empty(self):
        assert compare_values("", "PLATA", TYPE_STRING) is False
        assert compare_values("PLATA", "", TYPE_STRING) is False
        assert compare_values("", "", TYPE_STRING) is True

    def test_number_empty_vs_zero(self):
        # vacío != 0 (cuidado con falsos positivos)
        assert compare_values("", 0, TYPE_NUMBER) is False

    def test_number_invalid_string_falls_to_string_compare(self):
        # si no parsea, comparar como string
        assert compare_values("N/A", "n/a", TYPE_NUMBER) is True

    def test_label_set_raises(self):
        # _labels_ se intercepta en Task 12; aquí debe explotar
        with pytest.raises(ValueError, match="Unknown type_"):
            compare_values("A | B", "B | A", "label_set")

    def test_unknown_type_raises(self):
        with pytest.raises(ValueError, match="Unknown type_"):
            compare_values("x", "x", "bogus_type")


class TestComputeFieldDiffs:
    def _pair(self, xlsm_raw, sh_raw, xlsm_labels=None, sh_labels=None):
        x = _mk_row(source="xlsm", labels=xlsm_labels or ["", "", "", "", ""])
        s = _mk_row(id_sh="100", labels=sh_labels or ["", "", "", "", ""])
        x.raw = xlsm_raw
        s.raw = sh_raw
        return x, s

    def test_overwrite_when_different(self):
        x, s = self._pair({"Proceso": "PLATA"}, {"Proceso": "ZINC"})
        x.proceso = "PLATA"
        s.proceso = "ZINC"
        diffs = compute_field_diffs(x, s)
        assert {"field": "Proceso", "xlsm": "PLATA", "sh": "ZINC", "action": "overwrite"} in diffs

    def test_overwrite_when_sh_empty(self):
        x, s = self._pair({"Proceso": "PLATA"}, {"Proceso": ""})
        x.proceso = "PLATA"
        diffs = compute_field_diffs(x, s)
        assert any(d["field"] == "Proceso" for d in diffs)

    def test_no_diff_when_equal(self):
        x, s = self._pair({"Proceso": "PLATA"}, {"Proceso": "plata"})
        x.proceso = "PLATA"
        s.proceso = "plata"
        diffs = compute_field_diffs(x, s)
        assert all(d["field"] != "Proceso" for d in diffs)

    def test_conservative_does_not_overwrite(self):
        x, s = self._pair({"Descripción": "DESC X"}, {"Descripción": "DESC Y"})
        x.descripcion = "DESC X"
        s.descripcion = "DESC Y"
        diffs = compute_field_diffs(x, s)
        assert all(d["field"] != "Descripción" for d in diffs)

    def test_conservative_writes_when_sh_empty(self):
        x, s = self._pair({"Descripción": "DESC X"}, {"Descripción": ""})
        x.descripcion = "DESC X"
        diffs = compute_field_diffs(x, s)
        assert {"field": "Descripción", "xlsm": "DESC X", "sh": "", "action": "fill"} in diffs

    def test_empty_xlsm_never_writes(self):
        x, s = self._pair({"Proceso": ""}, {"Proceso": "ZINC"})
        diffs = compute_field_diffs(x, s)
        assert all(d["field"] != "Proceso" for d in diffs)

    def test_dash_propagates_for_overwrite(self):
        x, s = self._pair({"Proceso": "-"}, {"Proceso": "ZINC"})
        x.proceso = "-"
        s.proceso = "ZINC"
        diffs = compute_field_diffs(x, s)
        assert {"field": "Proceso", "xlsm": "-", "sh": "ZINC", "action": "delete"} in diffs

    def test_dash_in_conservative_does_not_propagate(self):
        x, s = self._pair({"Descripción": "-"}, {"Descripción": "DESC Y"})
        x.descripcion = "-"
        s.descripcion = "DESC Y"
        diffs = compute_field_diffs(x, s)
        assert all(d["field"] != "Descripción" for d in diffs)

    def test_labels_set_equal_no_diff(self):
        x, s = self._pair({}, {},
                          xlsm_labels=["PLATA", "ANTITARNISH", "", "", ""],
                          sh_labels=["ANTITARNISH", "PLATA", "", "", ""])
        diffs = compute_field_diffs(x, s)
        assert all(d["field"] != "_labels_" for d in diffs)

    def test_labels_set_xlsm_has_extra(self):
        x, s = self._pair({}, {},
                          xlsm_labels=["PLATA", "ANTITARNISH", "", "", ""],
                          sh_labels=["PLATA", "", "", "", ""])
        diffs = compute_field_diffs(x, s)
        label_diffs = [d for d in diffs if d["field"] == "_labels_"]
        assert len(label_diffs) == 1
        assert label_diffs[0]["xlsm_labels"] == ["PLATA", "ANTITARNISH", "", "", ""]
        assert label_diffs[0]["sh_extra"] == []

    def test_labels_set_sh_has_extra(self):
        x, s = self._pair({}, {},
                          xlsm_labels=["PLATA", "", "", "", ""],
                          sh_labels=["PLATA", "ZINC", "", "", ""])
        diffs = compute_field_diffs(x, s)
        label_diffs = [d for d in diffs if d["field"] == "_labels_"]
        assert len(label_diffs) == 1
        assert "ZINC" in label_diffs[0]["sh_extra"]

    def test_number_within_tolerance_no_diff(self):
        x, s = self._pair({"Plata (kg/pza)": 0.5000001}, {"Plata (kg/pza)": 0.5})
        diffs = compute_field_diffs(x, s)
        assert all(d["field"] != "Plata (kg/pza)" for d in diffs)


class TestEmitV11Xlsx:
    def test_emits_one_row_per_correction(self, tmp_path):
        template = "/Users/oviazcan/Projects/Ecoplating/SteelheadAutomator/remote/templates/Plantilla_Cotizaciones_v11.xlsm"
        out = tmp_path / "out.xlsm"

        corrections = [{
            "idSH": "12345",
            "customer": "SCHNEIDER",
            "pn": "ABC-001",
            "tier": "quoteIBMS",
            "diffs": [
                {"field": "Proceso", "xlsm": "PLATA SELECTIVA", "sh": "", "action": "fill"},
                {"field": "_labels_", "xlsm_labels": ["PLATA", "ANTITARNISH", "", "", ""],
                 "sh_labels": ["PLATA", "", "", "", ""], "sh_extra": [], "action": "overwrite"},
                {"field": "Plata (kg/pza)", "xlsm": "0.5", "sh": "", "action": "fill"},
            ],
        }]

        emit_v11_xlsx(template_path=template, corrections=corrections, out_path=out)

        wb = _openpyxl.load_workbook(out, read_only=True, data_only=True, keep_vba=True)
        ws = wb["Upload"]
        assert ws.cell(row=9, column=5).value in ("12345", 12345)
        assert ws.cell(row=9, column=6).value == "SCHNEIDER"
        assert ws.cell(row=9, column=7).value == "ABC-001"

    def test_omits_pns_without_diffs(self, tmp_path):
        template = "/Users/oviazcan/Projects/Ecoplating/SteelheadAutomator/remote/templates/Plantilla_Cotizaciones_v11.xlsm"
        out = tmp_path / "out.xlsm"
        emit_v11_xlsx(template_path=template, corrections=[], out_path=out)

        wb = _openpyxl.load_workbook(out, read_only=True, data_only=True, keep_vba=True)
        ws = wb["Upload"]
        assert ws.cell(row=9, column=5).value in (None, "")


class TestEmitJsonReport:
    def test_full_structure(self, tmp_path):
        out = tmp_path / "report.json"
        emit_json_report(
            out_path=out,
            inputs={"srg_xlsm": "a.xlsm", "cg_xlsm": "b.xlsm", "sh_report": "r.xlsx"},
            counts={
                "xlsm_rows_total": 100,
                "sh_rows_total": 200,
                "sh_rows_in_round": 110,
                "matched": 95,
                "suspicious_matches": 5,
                "unmatched_xlsm": 5,
                "unmatched_sh_in_round": 15,
                "duplicate_quoteibms_buckets": 2,
                "corrections_emitted": 80,
            },
            field_correction_counts={"Proceso": 10, "_labels_": 25},
            corrections=[{"idSH": "1", "customer": "C", "pn": "P", "tier": "quoteIBMS", "diffs": []}],
            suspicious_matches=[],
            unmatched_xlsm=[],
            unmatched_sh_in_round=[],
            duplicate_quoteibms=[],
        )
        data = json.loads(out.read_text())
        assert data["counts"]["matched"] == 95
        assert "generated_at" in data
        assert data["inputs"]["srg_xlsm"] == "a.xlsm"
        assert isinstance(data["corrections"], list)
