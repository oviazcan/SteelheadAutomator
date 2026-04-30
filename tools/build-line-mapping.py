#!/usr/bin/env python3
# Genera el mapping processName -> codigoLineaRecubrimiento ("T<n>") desde
# "1._Proceso - Tratamientos Genericos.xlsx", hoja "labels".
# Salida: imprime JSON por stdout y escribe remote/scripts/lib/process-line-mapping.json
# Uso: python3 tools/build-line-mapping.py

import json
import os
import re
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("ERROR: openpyxl no instalado. Instalalo con: pip3 install openpyxl")

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "1._Proceso - Tratamientos Genericos.xlsx")
OUT = os.path.join(ROOT, "remote", "scripts", "lib", "process-line-mapping.json")

# col index 0-based: A=0, B=1, C=2, D=3
COL_LINE_NAME = 2  # "Nombre de la Línea"
COL_PROCESS = 3    # "PROCESO"

LINE_RE = re.compile(r"^(T\w+?)-LI\b", re.IGNORECASE)

def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    if "labels" not in wb.sheetnames:
        sys.exit(f"ERROR: hoja 'labels' no encontrada en {XLSX}")
    ws = wb["labels"]
    mapping = {}
    skipped = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if len(row) <= COL_PROCESS:
            continue
        line_name = row[COL_LINE_NAME]
        process = row[COL_PROCESS]
        if not process or not line_name:
            continue
        process = str(process).strip()
        line_name = str(line_name).strip()
        m = LINE_RE.match(line_name)
        if not m:
            skipped.append((i, process, line_name))
            continue
        code = m.group(1).upper()
        prev = mapping.get(process)
        if prev and prev != code:
            print(f"WARN R{i}: '{process}' aparece con codigos distintos ({prev} vs {code}); conservo {prev}", file=sys.stderr)
            continue
        mapping[process] = code

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(mapping, f, ensure_ascii=False, indent=2, sort_keys=True)

    print(f"OK: {len(mapping)} procesos mapeados -> {OUT}", file=sys.stderr)
    if skipped:
        print(f"WARN: {len(skipped)} filas saltadas (sin patron T<n>-LI):", file=sys.stderr)
        for i, p, l in skipped[:10]:
            print(f"  R{i}: process='{p[:40]}' lineName='{l[:40]}'", file=sys.stderr)
    print(json.dumps(mapping, ensure_ascii=False, indent=2, sort_keys=True))

if __name__ == "__main__":
    main()
