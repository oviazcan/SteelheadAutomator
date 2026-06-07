#!/usr/bin/env python3
"""Dual-source offline recovery.

Cruza los xlsm originales de bulk-upload contra el reporte oficial de
Steelhead y emite un xlsx v11 con las correcciones por Id SH.

Uso:
    python3 tools/dual_source_recovery.py \\
        --srg-xlsm SRG.xlsm \\
        --cg-xlsm  CG.xlsm \\
        --sh-report report.xlsx \\
        --template remote/templates/Plantilla_Cotizaciones_v11.xlsm \\
        --out-xlsx recovery.xlsm \\
        --out-csv  recovery.csv      # opcional (v1.0.4): CSV directo para bulk-upload
        --report-json report.json

Spec: docs/superpowers/specs/2026-05-27-dual-source-offline-recovery-design.md
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import defaultdict
from dataclasses import dataclass, field, asdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

import openpyxl


def norm(value) -> str:
    """Normaliza una celda para comparación: cast a str, lower, strip, colapsa whitespace."""
    if value is None or value == "":
        return ""
    s = str(value)
    return re.sub(r"\s+", " ", s.strip().lower())


ROUND_MARKER_TOKENS = (
    "F1:", "F2:", "F3:", "F4:", "F5:",
    "MPO:", "SPECS:", "DEPT:", "METAL:", "PROC:",
    "LSPEC1:", "USPEC1:", "LSPEC2:", "USPEC2:",
    "SPECIALREQ:",
)


def is_round_marker(notas) -> bool:
    """True si las Notas adicionales contienen al menos 3 tokens estructurados
    de esta ronda, separados por ` | `."""
    if not notas:
        return False
    text = str(notas).upper()
    # contar tokens (case-insensitive)
    count = sum(1 for tok in ROUND_MARKER_TOKENS if tok in text)
    if count < 3:
        return False
    # exigir separador ` | ` al menos una vez (descarta freeform que tenga colons)
    return " | " in text


def make_fingerprint(metal_base: str | None, labels: Iterable[str | None]) -> str:
    """Fingerprint estable para discriminar PNs con mismo (cliente, nombre).

    Usa metalBase + labels normalizados y ordenados. Mismo formato que usa
    `audit-incomplete-pns.js` fase 5.4b para que los resultados sean comparables.
    """
    metal = (str(metal_base) if metal_base else "").strip().upper()
    labels_clean = sorted({
        str(l).strip().upper() for l in labels if l is not None and str(l).strip() != ""
    })
    return f"{metal}|{','.join(labels_clean)}"


def read_header_map(ws, header_row: int) -> dict[str, int]:
    """Devuelve {header_name → col_index (1-based)} leyendo la fila indicada.

    Normaliza headers: reemplaza '\\n' por espacio y colapsa whitespaces.
    """
    headers: dict[str, int] = {}
    for row in ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True):
        for idx, val in enumerate(row, start=1):
            if val is None or str(val).strip() == "":
                continue
            key = re.sub(r"\s+", " ", str(val).replace("\n", " ").strip())
            headers[key] = idx
        break
    return headers


@dataclass
class PartNumberRow:
    """Una fila de PN normalizada — fuente puede ser xlsm o reporte SH."""
    source: str  # 'xlsm_srg' | 'xlsm_cg' | 'sh_report'
    source_row: int  # fila 1-based del archivo origen (debug)
    id_sh: str  # vacío en xlsm; numérico-string en SH report
    cliente: str
    pn: str
    descripcion: str
    quote_ibms: str
    est_ibms: str
    notas: str
    metal_base: str
    labels: list[str]  # 5 elementos, "" para vacíos
    proceso: str
    spec1: str
    spec1_um: str
    spec2: str
    spec2_um: str
    # campos numéricos / sobreescribir
    raw: dict[str, object] = field(default_factory=dict)  # acceso por header_name → valor crudo

    def get(self, header: str):
        return self.raw.get(header)


# Headers requeridos en el reporte SH (los que el script lee SIEMPRE)
REQUIRED_SH_HEADERS = (
    "Id SH", "Cliente", "Número de parte", "Descripción",
    "Metal base", "Etiqueta 1", "Etiqueta 2", "Etiqueta 3", "Etiqueta 4", "Etiqueta 5",
    "Proceso", "Spec 1", "Spec 2", "Notas adicionales", "QuoteIBMS",
)


def load_sh_report(path: str | Path) -> list[PartNumberRow]:
    """Lee el reporte oficial de SH (xlsx, hoja única, headers row 1)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        headers = read_header_map(ws, header_row=1)

        for req in REQUIRED_SH_HEADERS:
            if req not in headers:
                raise ValueError(f"header esperado en reporte SH no encontrado: {req!r}")

        rows: list[PartNumberRow] = []
        for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # skip filas completamente vacías
            if all(c is None or str(c).strip() == "" for c in row):
                continue
            raw = {h: row[i - 1] if i - 1 < len(row) else None for h, i in headers.items()}
            pn_row = _row_from_raw(source="sh_report", source_row=r_idx, raw=raw)
            # skip si no tiene id ni pn (row con ruido)
            if not pn_row.id_sh and not pn_row.pn:
                continue
            rows.append(pn_row)
        return rows
    finally:
        wb.close()


def _row_from_raw(source: str, source_row: int, raw: dict) -> PartNumberRow:
    """Construye un PartNumberRow desde un dict {header → valor}."""
    def s(key: str) -> str:
        v = raw.get(key)
        if v is None:
            return ""
        return str(v).strip()

    labels = [s(f"Etiqueta {i}") for i in range(1, 6)]
    return PartNumberRow(
        source=source,
        source_row=source_row,
        id_sh=s("Id SH"),
        cliente=s("Cliente"),
        pn=s("Número de parte"),
        descripcion=s("Descripción"),
        quote_ibms=s("QuoteIBMS"),
        est_ibms=s("EstIBMS"),
        notas=s("Notas adicionales"),
        metal_base=s("Metal base"),
        labels=labels,
        proceso=s("Proceso"),
        spec1=s("Spec 1"),
        spec1_um=s("Esp. Spec 1 (µm)"),
        spec2=s("Spec 2"),
        spec2_um=s("Esp. Spec 2 (µm)"),
        raw=raw,
    )


REQUIRED_XLSM_HEADERS = (
    "Cliente", "Número de parte", "Metal base", "Proceso",
    "Etiqueta 1", "Etiqueta 2", "Etiqueta 3", "Etiqueta 4", "Etiqueta 5",
    "Spec 1", "Spec 2", "Notas adicionales", "QuoteIBMS",
)


def load_xlsm_originals(sources: list[tuple[str | Path, str]]) -> list[PartNumberRow]:
    """Lee uno o más xlsm de bulk-upload (hoja 'Upload', headers row 7, datos row 9+).

    `sources` es lista de (path, source_label).
    """
    all_rows: list[PartNumberRow] = []
    for path, source_label in sources:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_links=False)
        try:
            if "Upload" not in wb.sheetnames:
                raise ValueError(f"{path}: hoja 'Upload' no encontrada")
            ws = wb["Upload"]
            headers = read_header_map(ws, header_row=7)
            for req in REQUIRED_XLSM_HEADERS:
                if req not in headers:
                    raise ValueError(f"{path}: header esperado no encontrado: {req!r}")
            for r_idx, row in enumerate(ws.iter_rows(min_row=9, values_only=True), start=9):
                if all(c is None or str(c).strip() == "" for c in row):
                    continue
                raw = {h: row[i - 1] if i - 1 < len(row) else None for h, i in headers.items()}
                pn_row = _row_from_raw(source=source_label, source_row=r_idx, raw=raw)
                if not pn_row.pn:
                    continue
                all_rows.append(pn_row)
        finally:
            wb.close()
    return all_rows


def filter_round(rows: list[PartNumberRow]) -> list[PartNumberRow]:
    """Filtra rows que tienen la marca estructurada de 'esta ronda' en Notas adicionales."""
    return [r for r in rows if is_round_marker(r.notas)]


@dataclass
class MatchedPair:
    xlsm_row: PartNumberRow
    sh_row: PartNumberRow
    tier: str  # 'quoteIBMS' | 'composite_unique' | 'fingerprint'


@dataclass
class MatchResult:
    matched: list[MatchedPair] = field(default_factory=list)
    unmatched_xlsm: list[dict] = field(default_factory=list)
    unmatched_sh_in_round: list[dict] = field(default_factory=list)
    duplicate_quoteibms: list[dict] = field(default_factory=list)


def _key_composite(row: PartNumberRow) -> tuple[str, str]:
    return (row.cliente.strip().upper(), row.pn.strip().upper())


def match_xlsm_to_sh(xlsm_rows: list[PartNumberRow], sh_round: list[PartNumberRow]) -> MatchResult:
    # Index sh_round por QuoteIBMS y por (cliente, pn)
    by_quote: dict[str, list[PartNumberRow]] = defaultdict(list)
    by_key: dict[tuple[str, str], list[PartNumberRow]] = defaultdict(list)
    for sh in sh_round:
        if sh.quote_ibms:
            by_quote[sh.quote_ibms.strip().upper()].append(sh)
        by_key[_key_composite(sh)].append(sh)

    # Detectar duplicate_quoteibms buckets (≥2 candidatos)
    dup_qibms: list[dict] = []
    for q, cands in by_quote.items():
        if len(cands) >= 2:
            dup_qibms.append({
                "quoteIBMS": q,
                "candidates": [
                    {"idSH": c.id_sh, "pn": c.pn, "customer": c.cliente}
                    for c in cands
                ],
            })

    result = MatchResult(duplicate_quoteibms=dup_qibms)
    matched_sh_ids: set[str] = set()

    for x in xlsm_rows:
        # Tier 1: QuoteIBMS único
        if x.quote_ibms:
            q = x.quote_ibms.strip().upper()
            cands = by_quote.get(q, [])
            if len(cands) == 1:
                sh_match = cands[0]
                result.matched.append(MatchedPair(xlsm_row=x, sh_row=sh_match, tier="quoteIBMS"))
                matched_sh_ids.add(sh_match.id_sh)
                continue
            # 0 o ≥2 candidatos → cae a Tier 2 (ya registrado en dup_qibms si ≥2)

        # Tier 2: (cliente, pn) único
        key = _key_composite(x)
        cands = by_key.get(key, [])
        if len(cands) == 0:
            result.unmatched_xlsm.append({
                "pn": x.pn, "customer": x.cliente, "quoteIBMS": x.quote_ibms,
                "source": x.source, "source_row": x.source_row,
                "reason": "no_pn_in_sh_round",
            })
            continue
        if len(cands) == 1:
            result.matched.append(MatchedPair(xlsm_row=x, sh_row=cands[0], tier="composite_unique"))
            matched_sh_ids.add(cands[0].id_sh)
            continue

        # Tier 3: fingerprint
        x_fp = make_fingerprint(x.metal_base, x.labels)
        fp_matches = [c for c in cands if make_fingerprint(c.metal_base, c.labels) == x_fp]
        if len(fp_matches) == 1:
            result.matched.append(MatchedPair(xlsm_row=x, sh_row=fp_matches[0], tier="fingerprint"))
            matched_sh_ids.add(fp_matches[0].id_sh)
        else:
            result.unmatched_xlsm.append({
                "pn": x.pn, "customer": x.cliente, "quoteIBMS": x.quote_ibms,
                "source": x.source, "source_row": x.source_row,
                "reason": "ambiguous_fingerprint",
                "candidates": [{"idSH": c.id_sh, "fingerprint": make_fingerprint(c.metal_base, c.labels)} for c in cands],
            })

    # PNs en sh_round que no quedaron matcheados
    for sh in sh_round:
        if sh.id_sh not in matched_sh_ids:
            result.unmatched_sh_in_round.append({
                "idSH": sh.id_sh, "pn": sh.pn, "customer": sh.cliente,
                "notas": sh.notas[:200],
            })

    return result


# Acciones permitidas por columna
ACTION_NEVER = "never"
ACTION_CONSERVATIVE = "conservative"  # escribir solo si SH vacío
ACTION_OVERWRITE = "overwrite"        # escribir si difiere

# Tipos de comparación
TYPE_STRING = "string"
TYPE_NUMBER = "number"
TYPE_LABEL_SET = "label_set"  # cols 1-5 como set

FIELD_RULES: list[dict] = [
    # (header_name, action, type)
    {"header": "Descripción",                 "action": ACTION_CONSERVATIVE, "type": TYPE_STRING},
    {"header": "PN alterno",                  "action": ACTION_CONSERVATIVE, "type": TYPE_STRING},
    {"header": "Grupo",                       "action": ACTION_CONSERVATIVE, "type": TYPE_STRING},
    {"header": "Línea",                       "action": ACTION_CONSERVATIVE, "type": TYPE_STRING},
    # v1.0.3: Metal base es ACTION_OVERWRITE — el template v11 pre-llena la
    # celda con '(seleccione o escriba)'. Si la regla era CONSERVATIVE y SH ya
    # tenía Metal base, recovery no emitía nada → la celda quedaba con el
    # placeholder, lo que rompía la fórmula CNE del template al abrir en Excel.
    # BD es la fuente de verdad para Metal base (el usuario lo validó manualmente).
    {"header": "Metal base",                  "action": ACTION_OVERWRITE,    "type": TYPE_STRING},
    {"header": "_labels_",                    "action": ACTION_OVERWRITE,    "type": TYPE_LABEL_SET},  # especial: cols 1-5
    {"header": "Proceso",                     "action": ACTION_OVERWRITE,    "type": TYPE_STRING},
    {"header": "Spec 1",                      "action": ACTION_OVERWRITE,    "type": TYPE_STRING},
    {"header": "Esp. Spec 1 (µm)",            "action": ACTION_OVERWRITE,    "type": TYPE_NUMBER},
    {"header": "Spec 2",                      "action": ACTION_OVERWRITE,    "type": TYPE_STRING},
    {"header": "Esp. Spec 2 (µm)",            "action": ACTION_OVERWRITE,    "type": TYPE_NUMBER},
    {"header": "KGM (kg/pza)",                "action": ACTION_OVERWRITE,    "type": TYPE_NUMBER},
    {"header": "CMK (cm²/pza)",               "action": ACTION_OVERWRITE,    "type": TYPE_NUMBER},
    {"header": "LM (m/pza)",                  "action": ACTION_OVERWRITE,    "type": TYPE_NUMBER},
    {"header": "Mín Pzas Lote",               "action": ACTION_CONSERVATIVE, "type": TYPE_NUMBER},
    {"header": "Rack Flybar o Barril (Carga)","action": ACTION_CONSERVATIVE, "type": TYPE_STRING},
    {"header": "Pzas/Rack Línea",             "action": ACTION_CONSERVATIVE, "type": TYPE_NUMBER},
    {"header": "Rack Específico",             "action": ACTION_CONSERVATIVE, "type": TYPE_STRING},
    {"header": "Pzas/Rack Sec.",              "action": ACTION_CONSERVATIVE, "type": TYPE_NUMBER},
    {"header": "Tipo de Geometría",           "action": ACTION_CONSERVATIVE, "type": TYPE_STRING},
    {"header": "Longitud (m)",                "action": ACTION_CONSERVATIVE, "type": TYPE_NUMBER},
    {"header": "Ancho (m)",                   "action": ACTION_CONSERVATIVE, "type": TYPE_NUMBER},
    {"header": "Alto (m)",                    "action": ACTION_CONSERVATIVE, "type": TYPE_NUMBER},
    {"header": "Diám.Ext (m)",                "action": ACTION_CONSERVATIVE, "type": TYPE_NUMBER},
    {"header": "Diám.Int (m)",                "action": ACTION_CONSERVATIVE, "type": TYPE_NUMBER},
    {"header": "Departamento",                "action": ACTION_CONSERVATIVE, "type": TYPE_STRING},
    {"header": "Código SAT",                  "action": ACTION_CONSERVATIVE, "type": TYPE_STRING},
    {"header": "Plata (kg/pza)",              "action": ACTION_OVERWRITE,    "type": TYPE_NUMBER},
    {"header": "Estaño (kg/pza)",             "action": ACTION_OVERWRITE,    "type": TYPE_NUMBER},
    {"header": "Níquel (kg/pza)",             "action": ACTION_OVERWRITE,    "type": TYPE_NUMBER},
    {"header": "Zinc (kg/pza)",               "action": ACTION_OVERWRITE,    "type": TYPE_NUMBER},
    {"header": "Cobre (kg/pza)",              "action": ACTION_OVERWRITE,    "type": TYPE_NUMBER},
    {"header": "Antitarnish (L/pza)",         "action": ACTION_OVERWRITE,    "type": TYPE_NUMBER},
    {"header": "Epóx. MT (lb/pza)",           "action": ACTION_OVERWRITE,    "type": TYPE_NUMBER},
    {"header": "Epóx. BT (lb/pza)",           "action": ACTION_OVERWRITE,    "type": TYPE_NUMBER},
    {"header": "Epóx. MTR (lb/pza)",          "action": ACTION_OVERWRITE,    "type": TYPE_NUMBER},
    {"header": "QuoteIBMS",                   "action": ACTION_OVERWRITE,    "type": TYPE_STRING},
    {"header": "EstIBMS",                     "action": ACTION_OVERWRITE,    "type": TYPE_STRING},
    {"header": "Plano",                       "action": ACTION_OVERWRITE,    "type": TYPE_STRING},
    {"header": "Piezas por Carga",            "action": ACTION_OVERWRITE,    "type": TYPE_NUMBER},
    {"header": "Cargas por Hora",             "action": ACTION_OVERWRITE,    "type": TYPE_NUMBER},
    {"header": "Tiempo de Entrega",           "action": ACTION_OVERWRITE,    "type": TYPE_NUMBER},
    # Notas adicionales: NUNCA se escribe (es validador)
    # Cliente, Número de parte: NUNCA (llaves; van como referencia visual fuera de FIELD_RULES)
    # Cantidad, Precio, Unidad precio, Divisa, Precio default, Productos 1-3: NUNCA (sensibles)
]

NUMERIC_TOLERANCE = 1e-5


def _to_float_or_none(v) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


def compare_values(xlsm_val, sh_val, type_: str) -> bool:
    """True si son equivalentes según el tipo.

    NOTA: TYPE_LABEL_SET se maneja en compute_field_diffs (Task 12) antes
    de llegar aquí. Si llega, es un bug y explotamos.
    """
    if type_ == TYPE_NUMBER:
        fx = _to_float_or_none(xlsm_val)
        fs = _to_float_or_none(sh_val)
        if fx is None and fs is None:
            return norm(xlsm_val) == norm(sh_val)
        if fx is None or fs is None:
            # uno vacío, otro número → distintos (incluso si el número es 0)
            return False
        return abs(fx - fs) <= NUMERIC_TOLERANCE
    if type_ == TYPE_STRING:
        return norm(xlsm_val) == norm(sh_val)
    raise ValueError(f"Unknown type_ in compare_values: {type_!r}")


def _norm_labels(labels: list[str]) -> list[str]:
    """Limpia labels: strip por slot, preserva casing original y las 5 posiciones.

    No aplica `.upper()` — el bulk-upload v11 hace match case-sensitive contra el
    catálogo de Steelhead (`labelByName.has(name)`), que está en Title Case
    (`Plata`, `Estaño`, `Cromato Claro Azul (Iridiscente)`). Para comparaciones
    semánticas usar `_label_set` que sí normaliza a upper.
    """
    out: list[str] = []
    for lbl in (labels + ["", "", "", "", ""])[:5]:
        out.append(str(lbl or "").strip())
    return out


def _label_set(labels: list[str]) -> set[str]:
    """Set normalizado a upper para comparación case-insensitive."""
    return {lbl.upper() for lbl in _norm_labels(labels) if lbl}


def compute_field_diffs(xlsm_row: PartNumberRow, sh_row: PartNumberRow) -> list[dict]:
    """Genera la lista de campos a escribir según FIELD_RULES.

    Cada diff: {field, xlsm, sh, action}
      action ∈ {fill, overwrite, delete}
    Para labels: {field='_labels_', xlsm_labels:[...], sh_labels:[...], sh_extra:[...], action='overwrite'}
    """
    diffs: list[dict] = []

    for rule in FIELD_RULES:
        h = rule["header"]
        action = rule["action"]
        type_ = rule["type"]

        if h == "_labels_":
            x_labels = _norm_labels(xlsm_row.labels)
            s_labels = _norm_labels(sh_row.labels)
            x_set = _label_set(x_labels)
            s_set = _label_set(s_labels)
            if x_set == s_set:
                continue
            sh_extra = sorted(s_set - x_set)
            diffs.append({
                "field": "_labels_",
                "xlsm_labels": x_labels,
                "sh_labels": s_labels,
                "sh_extra": sh_extra,
                "action": "overwrite",
            })
            continue

        x_raw = xlsm_row.raw.get(h)
        s_raw = sh_row.raw.get(h)
        x_str = "" if x_raw is None else str(x_raw).strip()
        s_str = "" if s_raw is None else str(s_raw).strip()

        if x_str == "":
            continue

        if x_str == "-":
            if action == ACTION_OVERWRITE and s_str != "":
                diffs.append({"field": h, "xlsm": "-", "sh": s_str, "action": "delete"})
            continue

        equal = compare_values(x_raw, s_raw, type_)

        if action == ACTION_NEVER:
            continue
        if action == ACTION_CONSERVATIVE:
            if s_str == "":
                diffs.append({"field": h, "xlsm": x_str, "sh": "", "action": "fill"})
            continue
        if action == ACTION_OVERWRITE:
            if not equal:
                act = "overwrite" if s_str != "" else "fill"
                diffs.append({"field": h, "xlsm": x_str, "sh": s_str, "action": act})

    return diffs


def emit_v11_xlsx(template_path: str | Path, corrections: list[dict], out_path: str | Path) -> None:
    """Copia la plantilla v11 y agrega 1 fila por correction desde row 9.

    El template tiene macros (.xlsm). Usamos keep_vba=True para preservarlos.
    """
    wb = openpyxl.load_workbook(template_path, keep_vba=True, data_only=False)
    try:
        if "Upload" not in wb.sheetnames:
            raise ValueError(f"template no tiene hoja 'Upload': {template_path}")
        ws = wb["Upload"]
        headers = read_header_map(ws, header_row=7)

        required_v11 = ("Id SH", "Cliente", "Número de parte", "Proceso", "Notas adicionales")
        for r in required_v11:
            if r not in headers:
                raise ValueError(f"template v11: header esperado no encontrado: {r!r}")

        label_cols = [headers.get(f"Etiqueta {i}") for i in range(1, 6)]
        if any(c is None for c in label_cols):
            raise ValueError("template v11: Etiqueta 1-5 no resueltos")
        metal_base_col = headers.get("Metal base")
        if metal_base_col is None:
            raise ValueError("template v11: 'Metal base' no resuelto")
        # v1.0.5: PARÁMETROS (col A..D) del template traen V/F boilerplate en
        # las primeras ~500 filas (A='F', B='V', C='F', D='F'). En modo diff
        # NO queremos emitir estos toggles — bulk-upload los interpreta como
        # instrucciones explícitas. Incidente 2026-05-28: "F" en col Archivado
        # disparó STEP 8 unarchive masivo (10,842 PNs desarchivados sin
        # intención). Limpiar SIEMPRE.
        param_cols = []
        for name in ("Archivado", "Validación\n1er recibo", "Forzar\nduplicar", "Archivar\nanterior"):
            col = headers.get(name)
            if col is None:
                # Header con salto de línea puede venir como "Validación 1er recibo" según extractor; intentar fallback
                col = headers.get(name.replace("\n", " "))
            if col is None:
                raise ValueError(f"template v11: '{name}' (PARÁMETROS) no resuelto")
            param_cols.append(col)

        # Metadata row 3: Nombre Cotización/Layout (col 8) — facilita auditoría posterior
        ws.cell(row=3, column=8, value=f"Recovery dual-source {datetime.now(timezone.utc).strftime('%Y-%m-%d')}")

        start_row = 9
        for offset, corr in enumerate(corrections):
            r = start_row + offset
            ws.cell(row=r, column=headers["Id SH"], value=corr["idSH"])
            ws.cell(row=r, column=headers["Cliente"], value=corr["customer"])
            ws.cell(row=r, column=headers["Número de parte"], value=corr["pn"])
            # Limpiar columnas de etiqueta SIEMPRE: la plantilla v11 pre-llena
            # las celdas con '(seleccione)' y openpyxl no las sobreescribe cuando
            # value=None pasa por slots intermedios vacíos.
            for col in label_cols:
                ws.cell(row=r, column=col).value = None
            # v1.0.3: igual que las etiquetas, Metal base trae '(seleccione o escriba)'
            # como default — limpiar SIEMPRE para que sólo quede el valor del diff
            # (si lo hay) y para que las filas sin diff de Metal base no rompan
            # la fórmula CNE en Excel.
            ws.cell(row=r, column=metal_base_col).value = None
            # v1.0.5: limpiar PARÁMETROS A..D (Archivado, Validación 1er recibo,
            # Forzar duplicar, Archivar anterior). Ver comentario arriba.
            for col in param_cols:
                ws.cell(row=r, column=col).value = None

            for d in corr["diffs"]:
                if d["field"] == "_labels_":
                    xlsm_labels = d["xlsm_labels"]
                    for i, lab in enumerate(xlsm_labels, start=0):
                        if i >= 5:
                            break
                        col = label_cols[i]
                        if lab:
                            ws.cell(row=r, column=col, value=lab)
                    continue
                col = headers.get(d["field"])
                if col is None:
                    print(f"WARN: campo {d['field']!r} no existe en template v11, ignorado", file=sys.stderr)
                    continue
                value = d["xlsm"] if d["action"] != "delete" else "-"
                ws.cell(row=r, column=col, value=value)

        out_path = Path(out_path)
        wb.save(out_path)
    finally:
        wb.close()


_CSV_PLACEHOLDERS = {"(seleccione)", "(seleccione o escriba)", "seleccione", "seleccione o escriba"}


def _cell_to_csv(value) -> str:
    """Convierte un valor de celda openpyxl a string CSV.

    - None / placeholders → ''
    - floats integrales → 'N' (sin .0) para no romper QuoteIBMS u otros enteros.
    - strings → strip + drop placeholder.
    """
    if value is None:
        return ""
    if isinstance(value, str):
        s = value.strip()
        if s.lower() in _CSV_PLACEHOLDERS:
            return ""
        return s
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return repr(value)
    return str(value)


def emit_v11_csv(xlsm_path: str | Path, out_csv_path: str | Path) -> dict:
    """Convierte un v11.xlsm ya emitido a CSV listo para bulk-upload v11.

    Formato:
      - utf-8, separador ',', quote '"', CRLF (idéntico a Excel "Save As CSV").
      - Conserva las filas 1-8 de metadata/headers: bulk-upload's parseRows las
        salta sola via detección por strings ('PARÁMETROS', 'Archivado', 'V/F',
        'Texto', 'Número de parte').
      - Limpia placeholders '(seleccione)' / '(seleccione o escriba)' → '' (defensa
        en profundidad; emit_v11_xlsx ya los limpia en Etq 1-5 y Metal base).

    Devuelve {rows_written, data_rows} para logging del caller.
    """
    import csv  # local import: módulo stdlib, evita costo si no se usa CSV

    wb = openpyxl.load_workbook(xlsm_path, data_only=True, keep_vba=False)
    try:
        if "Upload" not in wb.sheetnames:
            raise ValueError(f"xlsm no tiene hoja 'Upload': {xlsm_path}")
        ws = wb["Upload"]
        out_p = Path(out_csv_path)
        out_p.parent.mkdir(parents=True, exist_ok=True)

        rows_written = 0
        data_rows = 0
        with open(out_p, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f, quoting=csv.QUOTE_MINIMAL, lineterminator="\r\n")
            for row in ws.iter_rows(values_only=True):
                out = [_cell_to_csv(v) for v in row]
                while out and out[-1] == "":
                    out.pop()
                w.writerow(out)
                rows_written += 1
                # heurística data row: col F (PN, índice 5) tiene valor y no es header
                if len(out) > 6 and out[6] and out[6] not in ("Texto", "Número de parte", "PN"):
                    pn_cell = out[6].strip()
                    if pn_cell:
                        data_rows += 1
    finally:
        wb.close()

    return {"rows_written": rows_written, "data_rows": data_rows}


def emit_json_report(
    out_path: str | Path,
    inputs: dict,
    counts: dict,
    field_correction_counts: dict,
    corrections: list,
    suspicious_matches: list,
    unmatched_xlsm: list,
    unmatched_sh_in_round: list,
    duplicate_quoteibms: list,
) -> None:
    payload = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "inputs": inputs,
        "counts": counts,
        "field_correction_counts": field_correction_counts,
        "corrections": corrections,
        "suspicious_matches": suspicious_matches,
        "unmatched_xlsm": unmatched_xlsm,
        "unmatched_sh_in_round": unmatched_sh_in_round,
        "duplicate_quoteibms": duplicate_quoteibms,
    }
    out_p = Path(out_path)
    out_p.parent.mkdir(parents=True, exist_ok=True)
    out_p.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


_X000D_RE = re.compile(r"_x000D_", flags=re.IGNORECASE)


def _notas_canon(value) -> str:
    """Canónico para comparar notas: norm() + quitar artifact `_x000D_` de Excel."""
    return norm(_X000D_RE.sub("", str(value or "")))


def validate_notas(xlsm_row: PartNumberRow, sh_row: PartNumberRow) -> str:
    """Validador único del match.

    Retorna 'ok' o 'suspicious'.
    'suspicious' SOLO si ambos lados tienen notas no vacías, no son idénticas
    tras `_notas_canon`, y NINGUNA es prefijo de la otra (cap distinto de Notas
    adicionales entre xlsm y SH no cuenta como discrepancia real).
    """
    nx = _notas_canon(xlsm_row.notas)
    ns = _notas_canon(sh_row.notas)
    if not nx or not ns or nx == ns:
        return "ok"
    if nx.startswith(ns) or ns.startswith(nx):
        return "ok"
    return "suspicious"


def _summarize_corrections(matched: list) -> tuple[list[dict], dict[str, int]]:
    """Aplica compute_field_diffs a cada match y genera (corrections[], field_counts)."""
    corrections: list[dict] = []
    field_counts: dict[str, int] = defaultdict(int)
    for pair in matched:
        diffs = compute_field_diffs(pair.xlsm_row, pair.sh_row)
        if not diffs:
            continue
        for d in diffs:
            field_counts[d["field"]] += 1
        corrections.append({
            "idSH": pair.sh_row.id_sh,
            "customer": pair.sh_row.cliente,
            "pn": pair.sh_row.pn,
            "tier": pair.tier,
            "diffs": diffs,
        })
    corrections.sort(key=lambda c: (c["customer"], c["pn"]))
    return corrections, dict(field_counts)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Dual-source offline recovery for Steelhead bulk-upload.")
    parser.add_argument("--srg-xlsm", required=True)
    parser.add_argument("--cg-xlsm", required=True)
    parser.add_argument("--sh-report", required=True)
    parser.add_argument("--template", required=True)
    parser.add_argument("--out-xlsx", required=True)
    parser.add_argument(
        "--out-csv",
        required=False,
        default=None,
        help="Si se pasa, emite también un CSV listo para bulk-upload v11 "
             "(carga directa, sin pasar por Excel).",
    )
    parser.add_argument("--report-json", required=True)
    args = parser.parse_args(argv)

    print("Loading xlsm originals...", file=sys.stderr)
    xlsm_rows = load_xlsm_originals([
        (args.srg_xlsm, "xlsm_srg"),
        (args.cg_xlsm, "xlsm_cg"),
    ])
    print(f"  xlsm rows total: {len(xlsm_rows)}", file=sys.stderr)

    print("Loading SH report...", file=sys.stderr)
    sh_rows = load_sh_report(args.sh_report)
    print(f"  sh rows total: {len(sh_rows)}", file=sys.stderr)

    sh_round = filter_round(sh_rows)
    print(f"  sh rows in round: {len(sh_round)}", file=sys.stderr)

    coverage = len(sh_round) / max(1, len(xlsm_rows))
    if coverage < 0.9:
        print(f"WARN: cobertura xlsm→sh_round={coverage:.0%} < 90%. Revisar regex de marker.", file=sys.stderr)

    print("Matching...", file=sys.stderr)
    match_result = match_xlsm_to_sh(xlsm_rows, sh_round)
    print(f"  matched: {len(match_result.matched)}", file=sys.stderr)
    print(f"  unmatched_xlsm: {len(match_result.unmatched_xlsm)}", file=sys.stderr)
    print(f"  unmatched_sh_in_round: {len(match_result.unmatched_sh_in_round)}", file=sys.stderr)
    print(f"  duplicate_quoteibms: {len(match_result.duplicate_quoteibms)}", file=sys.stderr)

    valid_matches: list = []
    suspicious: list[dict] = []
    for pair in match_result.matched:
        status = validate_notas(pair.xlsm_row, pair.sh_row)
        if status == "suspicious":
            suspicious.append({
                "idSH": pair.sh_row.id_sh,
                "customer": pair.sh_row.cliente,
                "pn": pair.sh_row.pn,
                "tier": pair.tier,
                "notas_xlsm": pair.xlsm_row.notas[:2000],
                "notas_sh": pair.sh_row.notas[:2000],
            })
        else:
            valid_matches.append(pair)
    print(f"  suspicious_matches: {len(suspicious)}", file=sys.stderr)

    print("Computing field diffs...", file=sys.stderr)
    corrections, field_counts = _summarize_corrections(valid_matches)
    print(f"  corrections_emitted: {len(corrections)}", file=sys.stderr)

    print(f"Writing xlsx → {args.out_xlsx}", file=sys.stderr)
    emit_v11_xlsx(template_path=args.template, corrections=corrections, out_path=args.out_xlsx)

    if args.out_csv:
        print(f"Writing csv → {args.out_csv}", file=sys.stderr)
        csv_stats = emit_v11_csv(xlsm_path=args.out_xlsx, out_csv_path=args.out_csv)
        print(f"  csv rows_written: {csv_stats['rows_written']} | data_rows: {csv_stats['data_rows']}", file=sys.stderr)

    print(f"Writing json → {args.report_json}", file=sys.stderr)
    counts = {
        "xlsm_rows_total": len(xlsm_rows),
        "sh_rows_total": len(sh_rows),
        "sh_rows_in_round": len(sh_round),
        "matched": len(match_result.matched),
        "suspicious_matches": len(suspicious),
        "unmatched_xlsm": len(match_result.unmatched_xlsm),
        "unmatched_sh_in_round": len(match_result.unmatched_sh_in_round),
        "duplicate_quoteibms_buckets": len(match_result.duplicate_quoteibms),
        "corrections_emitted": len(corrections),
    }
    emit_json_report(
        out_path=args.report_json,
        inputs={"srg_xlsm": args.srg_xlsm, "cg_xlsm": args.cg_xlsm, "sh_report": args.sh_report},
        counts=counts,
        field_correction_counts=field_counts,
        corrections=corrections,
        suspicious_matches=suspicious,
        unmatched_xlsm=match_result.unmatched_xlsm,
        unmatched_sh_in_round=match_result.unmatched_sh_in_round,
        duplicate_quoteibms=match_result.duplicate_quoteibms,
    )

    print(json.dumps(counts, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
