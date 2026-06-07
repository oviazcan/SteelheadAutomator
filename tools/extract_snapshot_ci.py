#!/usr/bin/env python3
"""extract_snapshot_ci.py

Extrae customInputs del snapshot SH (xlsx tipo "ING. Validación de Carga
Masiva NP 2026-05-27 (1).xlsx") a un JSON slim indexado por Id SH. El JSON
luego se carga en el script DevTools `recover-ci-from-snapshot.js` que hace
diff vs SH actual y emite SavePartNumber sólo para los que perdieron CI.

Output shape (array):
    [
      {
        "idsh": 3024889,
        "pn": "46004-157-01",
        "customer": "SCHNEIDER ELECTRIC MEXICO ...",
        "customInputs": {
          "DatosAdicionalesNP": {"BaseMetal": "Cobre", "QuoteIBMS": "58545", "EstacionIBMS": "58545"},
          "DatosFacturacion": {"CodigoSAT": "73181106 - ..."},
          "DatosPlanificacion": {"PiezasCarga": 4, "CargasHora": 12, "TiempoEntrega": 8},
          "NotasAdicionales": "F1: ..."
        }
      },
      ...
    ]

Nota: el shape coincide con el que SH guarda en `partNumber.customInputs`
(verificado contra el dryrun de recover-pair-update.js — QuoteIBMS/EstIBMS
viven en DatosAdicionalesNP, NO top-level; PiezasCarga/CargasHora sin "Por").

Uso:
  python3 tools/extract_snapshot_ci.py \\
      --xlsx "/Users/oviazcan/Downloads/ING. Validación de Carga Masiva NP 2026-05-27 (1).xlsx" \\
      --out tools/snapshot_ci_2026-05-27.json
"""
import argparse
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Falta openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

SHEET_NAME = "ING. Validación de Carga Masiva"

# Cols 1-based (verificadas en headers row 1)
COL_IDSH       = 5
COL_CLIENT     = 6
COL_PN         = 7
COL_METAL      = 17
COL_SAT        = 55
COL_NOTAS      = 65
COL_IBMS       = 66
COL_EST_IBMS   = 67
COL_PLANO      = 68
COL_PZAS       = 69
COL_CARGAS     = 70
COL_TE         = 71


def cell(row, c):
    idx = c - 1
    if idx >= len(row):
        return None
    return row[idx]


def s(v):
    if v is None:
        return None
    if isinstance(v, str):
        t = v.strip()
        return t if t else None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip() or None


def n(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        if isinstance(v, float) and v.is_integer():
            return int(v)
        return v
    if isinstance(v, str):
        t = v.strip()
        if not t:
            return None
        try:
            f = float(t)
            return int(f) if f.is_integer() else f
        except ValueError:
            return None
    return None


def build_ci(row):
    ci = {}
    adic = {}
    base = s(cell(row, COL_METAL))
    if base:
        adic["BaseMetal"] = base
    ibms = s(cell(row, COL_IBMS))
    if ibms:
        adic["QuoteIBMS"] = ibms
    est = s(cell(row, COL_EST_IBMS))
    if est:
        adic["EstacionIBMS"] = est
    plano = s(cell(row, COL_PLANO))
    if plano:
        adic["Plano"] = plano
    if adic:
        ci["DatosAdicionalesNP"] = adic

    sat = s(cell(row, COL_SAT))
    if sat:
        ci["DatosFacturacion"] = {"CodigoSAT": sat}

    plan = {}
    pcarga = n(cell(row, COL_PZAS))
    if pcarga is not None:
        plan["PiezasCarga"] = pcarga
    cph = n(cell(row, COL_CARGAS))
    if cph is not None:
        plan["CargasHora"] = cph
    te = n(cell(row, COL_TE))
    if te is None:
        te = s(cell(row, COL_TE))
    if te is not None and te != "":
        plan["TiempoEntrega"] = te
    if plan:
        ci["DatosPlanificacion"] = plan

    notas = s(cell(row, COL_NOTAS))
    if notas:
        ci["NotasAdicionales"] = notas

    return ci if ci else None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    p = Path(args.xlsx)
    if not p.exists():
        print(f"No existe: {p}", file=sys.stderr); sys.exit(2)

    wb = openpyxl.load_workbook(p, read_only=True, data_only=True, keep_links=False)
    if SHEET_NAME not in wb.sheetnames:
        print(f"Sin hoja '{SHEET_NAME}'. Hojas: {wb.sheetnames}", file=sys.stderr); sys.exit(2)
    ws = wb[SHEET_NAME]

    out = []
    seen_idsh = set()
    dups = 0
    no_idsh = 0
    no_ci = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        idsh_raw = cell(row, COL_IDSH)
        try:
            idsh = int(idsh_raw) if idsh_raw is not None else None
        except (ValueError, TypeError):
            idsh = None
        if idsh is None:
            no_idsh += 1
            continue
        if idsh in seen_idsh:
            dups += 1
            continue
        seen_idsh.add(idsh)

        ci = build_ci(row)
        if not ci:
            no_ci += 1
            # Lo guardamos igual con customInputs={} para que el diff sepa
            # que la fila existe pero está vacía (no es "missing from snapshot").
            ci = {}

        out.append({
            "idsh": idsh,
            "pn": s(cell(row, COL_PN)),
            "customer": s(cell(row, COL_CLIENT)),
            "customInputs": ci,
        })

    Path(args.out).write_text(json.dumps(out, ensure_ascii=False, separators=(',', ':')))

    with_ci = sum(1 for r in out if r["customInputs"])
    print(f"OK {len(out)} filas extraidas -> {args.out}")
    print(f"  Con algun CI:    {with_ci} ({100*with_ci/len(out):.1f}%)")
    print(f"  CI vacio:        {len(out)-with_ci}")
    print(f"  Saltadas sin id: {no_idsh}")
    print(f"  Duplicadas (id): {dups}")

    # Cobertura por campo
    counts = {"BaseMetal":0, "QuoteIBMS":0, "EstacionIBMS":0, "Plano":0,
              "CodigoSAT":0, "PiezasCarga":0, "CargasHora":0,
              "TiempoEntrega":0, "NotasAdicionales":0}
    for r in out:
        ci = r["customInputs"]
        da = ci.get("DatosAdicionalesNP", {})
        for k in ["BaseMetal","QuoteIBMS","EstacionIBMS","Plano"]:
            if da.get(k): counts[k] += 1
        if ci.get("DatosFacturacion",{}).get("CodigoSAT"): counts["CodigoSAT"] += 1
        dp = ci.get("DatosPlanificacion", {})
        for k in ["PiezasCarga","CargasHora","TiempoEntrega"]:
            if dp.get(k) is not None: counts[k] += 1
        if ci.get("NotasAdicionales"): counts["NotasAdicionales"] += 1
    print("Cobertura por campo:")
    for k, v in counts.items():
        print(f"  {k:18s}: {v}")


if __name__ == "__main__":
    main()
