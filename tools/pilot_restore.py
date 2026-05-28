#!/usr/bin/env python3
"""Reconstruye xlsm de RESTAURACIÓN para PNs del piloto, con filas COMPLETAS del xlsm fuente.

Contexto:
  El piloto string-only (50 PNs vía pilot_select.py) emitió xlsm con SOLO las
  columnas de diffs. El applet bulk-upload v11 hace REPLACE de customInputs
  (no merge), entonces los campos no-incluidos (Metal Base, Quote IBMS,
  Est IBMS, Plano cuando no era corrección, Piezas/Carga, Cargas/Hora,
  Lead Time, Notas Adicionales) llegaron como `null` → SH los borró.

  Este script repara el daño: para los mismos 50 PNs piloto, genera un xlsm
  con TODAS las columnas pobladas desde el xlsm fuente (las 2 BD Reloaded).
  Al cargar este xlsm, el applet construye un customInputs COMPLETO y SH lo
  reemplaza con los valores correctos.

Uso típico:
  python3 tools/pilot_restore.py \
    --report-json ~/Downloads/dualsource_report_v102.json \
    --srg-xlsm "$HOME/Downloads/BD Numeros de Parte Reloaded v23 valores base final SRG.xlsm" \
    --cg-xlsm "$HOME/Downloads/BD Numeros de Parte Reloaded v23 valores base final clientes generales.xlsm" \
    --template remote/templates/Plantilla_Cotizaciones_v11.xlsm \
    --out-xlsx ~/Downloads/recovery_pilot_50_RESTORE.xlsm \
    --limit 50 --max-per-customer 8
"""

from __future__ import annotations

import argparse
import json
import sys
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from dual_source_recovery import load_xlsm_originals, read_header_map  # noqa: E402
from pilot_select import select  # noqa: E402


SPEC_NAME_HEADERS = ("Spec 1", "Spec 2")
SPEC_ESP_HEADERS = ("Esp. Spec 1 (µm)", "Esp. Spec 2 (µm)")

# Headers de la plantilla v11 cuyo nombre NO matchea el del xlsm fuente.
# Detectado en validación piloto 50/50 PNs: la plantilla escribe "EstIBMS" mientras
# el xlsm fuente trae "EstacionIBMS". Sin el alias el campo quedaba None → SH lo
# borraba en customInputs.
TEMPLATE_TO_XLSM_HEADER_ALIASES: dict[str, str] = {
    "EstIBMS": "EstacionIBMS",
}


def _xlsm_value(raw: dict, template_header: str):
    """Lee el valor del xlsm fuente para un header de la plantilla v11.

    Aplica los aliases conocidos (`TEMPLATE_TO_XLSM_HEADER_ALIASES`). Si el
    header existe tal cual en el xlsm fuente lo devuelve; si no, prueba el alias.
    """
    if template_header in raw:
        return raw.get(template_header)
    alias = TEMPLATE_TO_XLSM_HEADER_ALIASES.get(template_header)
    if alias:
        return raw.get(alias)
    return raw.get(template_header)


def _is_dash_sentinel(value) -> bool:
    """`-` literal en el xlsm fuente = archive sentinel destructivo del bulk-upload.
    Cuando hacemos MODIFY full-row no queremos propagarlo: preferimos que el campo
    quede vacío y SH conserve lo que ya tenía.
    """
    return isinstance(value, str) and value.strip() == "-"


def _first_nonempty(values):
    for v in values:
        if v is None:
            continue
        if isinstance(v, str):
            s = v.strip()
            if not s or _is_dash_sentinel(s):
                continue
            return s
        return v
    return None


def _merge_specs(src_rows: list) -> tuple[list[str], list]:
    """Combina Spec 1+Spec 2 de varias filas en una lista única, preservando
    su espesor pareado. Filas legacy guardan una spec por step → al colapsar
    el PN duplicado en una sola fila SH, queremos ambas specs ligadas.
    """
    seen: set[str] = set()
    pairs: list[tuple[str, object]] = []
    for r in src_rows:
        for name_h, esp_h in zip(SPEC_NAME_HEADERS, SPEC_ESP_HEADERS):
            raw = r.raw.get(name_h)
            if not isinstance(raw, str):
                continue
            name = raw.strip()
            if not name or _is_dash_sentinel(name):
                continue
            if name in seen:
                continue
            seen.add(name)
            esp = r.raw.get(esp_h)
            pairs.append((name, esp))
    names = [p[0] for p in pairs]
    esps = [p[1] for p in pairs]
    return names, esps


def emit_v11_xlsx_full_rows(
    template_path: str | Path,
    rows_with_idsh: list[tuple[str, list, str]],  # (idSH, [src_rows], customer_for_label)
    out_path: str | Path,
) -> None:
    """Copia la plantilla v11 y escribe la fila del xlsm fuente por header name.

    Cuando hay varias filas xlsm para el mismo (pn, customer) — caso PNs Bimetales
    que en SH se consolidan en un solo PN con varias specs activas — las mergea:
    specs combinadas en Spec 1/Spec 2, demás campos toman el primer valor no-vacío.

    Cuando hay UNA sola fila, se copia tal cual (incluyendo `-` legítimos que el
    bulk-upload interpreta como archive sentinel "borrar las linked specs que no
    sean Spec 1").
    """
    wb = openpyxl.load_workbook(template_path, keep_vba=True, data_only=False)
    try:
        if "Upload" not in wb.sheetnames:
            raise ValueError(f"template no tiene hoja 'Upload': {template_path}")
        ws = wb["Upload"]
        headers = read_header_map(ws, header_row=7)

        if "Id SH" not in headers:
            raise ValueError("template v11: header 'Id SH' no encontrado")

        ws.cell(row=3, column=8, value=f"Restore FULL ROWS {datetime.now(timezone.utc).strftime('%Y-%m-%d')}")

        start_row = 9
        for offset, (idsh, src_rows, customer) in enumerate(rows_with_idsh):
            r = start_row + offset
            ws.cell(row=r, column=headers["Id SH"], value=idsh)
            if "Cliente" in headers:
                ws.cell(row=r, column=headers["Cliente"], value=customer)

            merge = len(src_rows) > 1
            spec_names, spec_esps = (_merge_specs(src_rows) if merge else (None, None))

            for header, col in headers.items():
                if header in ("Id SH", "Cliente"):
                    continue
                cell = ws.cell(row=r, column=col)

                if merge and header in SPEC_NAME_HEADERS:
                    idx = SPEC_NAME_HEADERS.index(header)
                    cell.value = spec_names[idx] if idx < len(spec_names) else None
                    continue
                if merge and header in SPEC_ESP_HEADERS:
                    idx = SPEC_ESP_HEADERS.index(header)
                    esp = spec_esps[idx] if idx < len(spec_esps) else None
                    if isinstance(esp, str):
                        esp = esp.strip() or None
                    cell.value = esp
                    continue

                if merge:
                    value = _first_nonempty(_xlsm_value(r2.raw, header) for r2 in src_rows)
                    cell.value = value
                    continue

                # Una sola fila: literal (preserva `-` legítimo del xlsm fuente).
                value = _xlsm_value(src_rows[0].raw, header)
                if isinstance(value, str):
                    s = value.strip()
                    cell.value = s if s else None
                else:
                    cell.value = value

        out_path = Path(out_path)
        wb.save(out_path)
    finally:
        wb.close()


def _resolve_xlsm_rows(
    pn: str,
    customer: str,
    by_pn: dict[str, list],
) -> tuple:
    """Encuentra TODAS las filas xlsm fuente para (pn, customer). Retorna (rows, warn|None)."""
    candidates = by_pn.get(pn, [])
    if not candidates:
        return None, f"sin filas xlsm para pn={pn!r}"
    if len(candidates) == 1:
        return list(candidates), None
    client_prefix = customer.split("—")[0].strip().upper()
    if not client_prefix:
        return list(candidates), None
    filtered = [c for c in candidates if (c.cliente or "").upper().startswith(client_prefix[:30])]
    if filtered:
        warn = None if len(filtered) == 1 else f"{len(filtered)} filas xlsm para pn={pn!r}, customer={customer!r}; combinando"
        return filtered, warn
    return list(candidates), f"customer={customer!r} no matchea exactamente; usando {len(candidates)} filas"


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    ap.add_argument("--report-json", required=True, type=Path)
    ap.add_argument("--srg-xlsm", required=True, type=Path)
    ap.add_argument("--cg-xlsm", required=True, type=Path)
    ap.add_argument("--template", required=True, type=Path)
    ap.add_argument("--out-xlsx", required=True, type=Path)
    ap.add_argument("--limit", type=int, default=50)
    ap.add_argument("--max-per-customer", type=int, default=8)
    args = ap.parse_args(argv)

    report = json.loads(args.report_json.read_text(encoding="utf-8"))
    corrections = report["corrections"]
    print(f"Loaded {len(corrections)} corrections from {args.report_json}", file=sys.stderr)

    picked = select(
        corrections,
        limit=args.limit,
        max_per_customer=args.max_per_customer,
    )
    print(f"Selected {len(picked)} PNs (same selector as pilot_select)", file=sys.stderr)

    print("Loading source xlsm files...", file=sys.stderr)
    xlsm_rows = load_xlsm_originals([
        (args.srg_xlsm, "xlsm_srg"),
        (args.cg_xlsm, "xlsm_cg"),
    ])
    print(f"  Loaded {len(xlsm_rows)} xlsm rows", file=sys.stderr)

    by_pn: dict[str, list] = defaultdict(list)
    for r in xlsm_rows:
        if r.pn:
            by_pn[r.pn].append(r)

    rows_with_idsh: list[tuple[str, list, str]] = []
    warnings: list[str] = []
    for p in picked:
        src_rows, warn = _resolve_xlsm_rows(p["pn"], p["customer"], by_pn)
        if src_rows is None:
            warnings.append(f"  SKIP pn={p['pn']!r} idSH={p['idSH']!r}: {warn}")
            continue
        if warn:
            warnings.append(f"  WARN pn={p['pn']!r} idSH={p['idSH']!r}: {warn}")
        rows_with_idsh.append((p["idSH"], src_rows, p["customer"]))

    print(f"Resolved {len(rows_with_idsh)}/{len(picked)} source rows", file=sys.stderr)
    if warnings:
        print("\nResolution warnings:", file=sys.stderr)
        for w in warnings:
            print(w, file=sys.stderr)

    emit_v11_xlsx_full_rows(
        template_path=str(args.template),
        rows_with_idsh=rows_with_idsh,
        out_path=str(args.out_xlsx),
    )
    print(f"\nWrote {args.out_xlsx}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
